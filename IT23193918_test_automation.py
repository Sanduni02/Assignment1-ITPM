import argparse
import time
import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

def parse_args():
    parser = argparse.ArgumentParser(description="Singlish Chat Translator Test Automation")
    parser.add_argument("--excel", required=True, help="Path to the Excel test cases file")
    parser.add_argument("--url", required=True, help="URL of the chat translator")
    parser.add_argument("--wait-ms", type=int, default=5000, help="Wait time in ms after submitting input")
    parser.add_argument("--type-delay-ms", type=int, default=80, help="Delay between keystrokes in ms")
    parser.add_argument("--slow-mo-ms", type=int, default=200, help="Slow motion delay for Playwright in ms")
    parser.add_argument("--save-every", type=int, default=1, help="Save Excel every N test cases")
    parser.add_argument("--keep-open", action="store_true", help="Keep browser open after all tests")
    return parser.parse_args()

def normalize(text):
    return " ".join(str(text).strip().split())

def find_input(page):
    selectors = [
        "textarea",
        "input[type='text']",
        "[contenteditable='true']",
        "[placeholder*='type']",
        "[placeholder*='Type']",
        "[placeholder*='enter']",
        "[placeholder*='Enter']",
        "[placeholder*='message']",
        "[placeholder*='input']",
        ".chat-input",
        ".input-box",
        "#chat-input",
        "#input",
        "form input",
        "form textarea",
    ]
    for sel in selectors:
        try:
            el = page.locator(sel).first
            el.wait_for(state="visible", timeout=3000)
            return el
        except Exception:
            continue
    return None

def find_output(page):
    selectors = [
        ".output",
        ".result",
        ".translation",
        ".translated",
        ".chat-output",
        ".response",
        "[class*='output']",
        "[class*='result']",
        "[class*='translat']",
        "[class*='response']",
        ".message:last-child",
        ".chat-message:last-child",
        ".bot-message:last-child",
        "p:last-of-type",
    ]
    for sel in selectors:
        try:
            els = page.locator(sel)
            count = els.count()
            if count > 0:
                text = els.last.inner_text().strip()
                if text:
                    return text
        except Exception:
            continue
    return ""

def clear_and_type(page, input_el, text, type_delay_ms):
    input_el.click()
    try:
        input_el.fill("")
    except Exception:
        input_el.triple_click()
        page.keyboard.press("Control+a")
        page.keyboard.press("Delete")
    page.keyboard.type(text, delay=type_delay_ms)

def submit_input(page, input_el):
    input_el.press("Enter")
    for btn_sel in ["button[type='submit']", "button:has-text('Send')", "button:has-text('Translate')", ".send-btn", "#send"]:
        try:
            btn = page.locator(btn_sel).first
            btn.wait_for(state="visible", timeout=1000)
            btn.click()
            return
        except Exception:
            continue

def run_tests(args):
    wb = openpyxl.load_workbook(args.excel)
    ws = wb.active

    col_map = {}
    for cell in ws[1]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    input_col    = col_map.get("Input")
    expected_col = col_map.get("Expected output")
    actual_col   = col_map.get("Actual output")
    status_col   = col_map.get("Status")

    if not all([input_col, expected_col, actual_col, status_col]):
        raise ValueError(f"Could not find required columns. Found: {list(col_map.keys())}")

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=False,
            slow_mo=args.slow_mo_ms,
            args=["--start-maximized"]
        )
        context = browser.new_context(no_viewport=True)
        page = context.new_page()

        print(f"\n{'='*60}")
        print(f"Navigating to: {args.url}")
        print(f"{'='*60}\n")

        try:
            page.goto(args.url, wait_until="networkidle", timeout=30000)
        except PlaywrightTimeoutError:
            page.goto(args.url, wait_until="load", timeout=30000)
        time.sleep(3)

        test_count = 0
        pass_count = 0
        fail_count = 0
        error_count = 0

        for row_idx in range(2, ws.max_row + 1):
            input_val    = ws.cell(row=row_idx, column=input_col).value
            expected_val = ws.cell(row=row_idx, column=expected_col).value

            if not input_val:
                continue

            input_text    = str(input_val).strip()
            expected_text = str(expected_val).strip() if expected_val else ""
            tc_id         = ws.cell(row=row_idx, column=1).value or f"Row {row_idx}"

            print(f"[{tc_id}] Input: {input_text[:70]}{'...' if len(input_text)>70 else ''}")

            try:
                input_el = find_input(page)
                if not input_el:
                    raise Exception("Could not locate an input element on the page.")

                clear_and_type(page, input_el, input_text, args.type_delay_ms)
                time.sleep(0.3)
                submit_input(page, input_el)
                time.sleep(args.wait_ms / 1000)

                actual_text = find_output(page)
                actual_norm   = normalize(actual_text)
                expected_norm = normalize(expected_text)
                status = "Pass" if actual_norm == expected_norm else "Fail"

                if status == "Pass":
                    pass_count += 1
                else:
                    fail_count += 1

                ws.cell(row=row_idx, column=actual_col).value = actual_text
                ws.cell(row=row_idx, column=status_col).value = status
                print(f"         Status : {status}")
                if status == "Fail":
                    print(f"         Expected: {expected_text[:60]}")
                    print(f"         Actual  : {actual_text[:60]}")

            except Exception as e:
                error_count += 1
                err_msg = f"ERROR: {e}"
                print(f"         *** {err_msg}")
                ws.cell(row=row_idx, column=actual_col).value = err_msg
                ws.cell(row=row_idx, column=status_col).value = "Error"

            test_count += 1
            if test_count % args.save_every == 0:
                wb.save(args.excel)

        wb.save(args.excel)

        print(f"\n{'='*60}")
        print(f"Test run complete.")
        print(f"  Total  : {test_count}")
        print(f"  Pass   : {pass_count}")
        print(f"  Fail   : {fail_count}")
        print(f"  Errors : {error_count}")
        print(f"Results saved to: {args.excel}")
        print(f"{'='*60}\n")

        if not args.keep_open:
            browser.close()
        else:
            print("Browser kept open. Press Enter here to close it.")
            input()
            browser.close()

if __name__ == "__main__":
    args = parse_args()
    run_tests(args)
